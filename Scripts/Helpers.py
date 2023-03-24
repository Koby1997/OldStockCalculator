import sys
from openpyxl import Workbook, load_workbook




"""
Gathers the user input

Parameters:
None

Returns:
workbook_path - path to the Excel sheet we are saving data to.
beginning_buy_range - 
end_buy_range - 
beginning_sell_range - 
end_sell_range - 

"""
def Collect_User_Input():

    workbook_name = input("What is the name of the Excel file you would like to save data to? (withought the .xlsx):   ")
    workbook_path = f"../Excel_Sheets/{workbook_name}.xlsx"

    print(  """\n\nEnter the ranges of buy and sell dates that you want to compute.
        The number you input is the number of days BEFORE the stock's Ex-dividend date.
        You can do negative numbers to act as days after the Ex-dividend date\n
        EXAMPLE:
        
        Beginning of Buy Date range: 21
        End of Buy Date range: 7
        Beginning of Sell Date range: 3
        End of Sell Date range: -3

        If you want to calculate the averages for all possibilities between buying 21 days before up to 7 days before
        and for selling 3 days before to 3 days after, you would input the values above.
        Now input your wanted values\n\n
        """)
    

    beginning_buy_range = int(input("Beginning of Buy Date range: "))
    end_buy_range = int(input("End of Buy Date range: "))
    beginning_sell_range = int(input("Beginning of Sell Date range: "))
    end_sell_range = int(input("End of Sell Date range: "))

    # Do logic checks now to make sure date ranges work

    # Just switch dates if the beginning/end date is further in the past than the beginning
    if end_buy_range > beginning_buy_range:
        end_buy_range, beginning_buy_range = beginning_buy_range, end_buy_range

    if end_sell_range > beginning_sell_range:
        end_sell_range, beginning_sell_range = beginning_sell_range, end_buy_range

    # Check if we are selling before we buy
    if beginning_sell_range >= end_buy_range:
        print("ERROR: Trying to sell before buying")
        print("Exiting Script")
        sys.exit()


    return (workbook_path,
            beginning_buy_range,
            end_buy_range,
            beginning_sell_range,
            end_sell_range)



























"""
Adds to RAW_DATA Sheet in Excel

Parameters:
workbook - current workbook we want to save data to
stock - the stock we currently have data for

Returns:
Nothing
"""
def Create_Excel_Raw_Data(workbook, stock, num_cols, num_rows):

    # Check if worksheet is already made, if not, make it
    if "RAW_DATA" in workbook.sheetnames:
        worksheet = workbook["RAW_DATA"]
    else:
        worksheet = workbook.create_sheet("RAW_DATA")
        worksheet.cell(row=1, column=1, value="Stock Symbol")
        worksheet.cell(row=2, column=1, value="Calculation Type")
        worksheet.cell(row=3, column=1, value="Data")
        worksheet.cell(row=6, column=1, value="Number of Columns")
        worksheet.cell(row=7, column=1, value=num_cols)
        worksheet.cell(row=9, column=1, value="Number of Rows")
        worksheet.cell(row=10, column=1, value=num_rows)

    # What is the next empty Column to place the new stock data
    col = Get_Next_Empty_Cell(worksheet)


    # Average Percent Change
    worksheet.cell(row=1, column=col, value=stock.symbol)
    worksheet.cell(row=2, column=col, value="Average Percent Change")

    row = 3
    for value in stock.multiple_perc_change:
        worksheet.cell(row=row, column=col, value=value)
        row += 1

    # High Outliers
    col += 1
    worksheet.cell(row=1, column=col, value=stock.symbol)
    worksheet.cell(row=2, column=col, value="High Outliers")

    row = 3
    for value in stock.high_perc_outliers:
        worksheet.cell(row=row, column=col, value=value)
        row += 1

    # Low Outliers
    col += 1
    worksheet.cell(row=1, column=col, value=stock.symbol)
    worksheet.cell(row=2, column=col, value="Low Outliers")

    row = 3
    for value in stock.low_perc_outliers:
        worksheet.cell(row=row, column=col, value=value)
        row += 1

    # Standard Deviation
    col += 1
    worksheet.cell(row=1, column=col, value=stock.symbol)
    worksheet.cell(row=2, column=col, value="Standard Deviation")

    row = 3
    for value in stock.multiple_SD:
        worksheet.cell(row=row, column=col, value=value)
        row += 1

    # Skewness
    col += 1
    worksheet.cell(row=1, column=col, value=stock.symbol)
    worksheet.cell(row=2, column=col, value="Skewness")

    row = 3
    for value in stock.multiple_skew:
        worksheet.cell(row=row, column=col, value=value)
        row += 1

    # Kurtosis
    col += 1
    worksheet.cell(row=1, column=col, value=stock.symbol)
    worksheet.cell(row=2, column=col, value="Kurtosis")

    row = 3
    for value in stock.multiple_kurt:
        worksheet.cell(row=row, column=col, value=value)
        row += 1

    





def Get_Next_Empty_Cell(worksheet):
    for cell in worksheet[1]:
        if cell.value is None:
                return cell.column
    return worksheet.max_column + 1