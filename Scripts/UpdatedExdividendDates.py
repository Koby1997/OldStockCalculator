import yfinance as yf
import datetime
import pandas as pd
import os
import sys
from StockClass import Stock
import Helpers
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# TODO try to make bell curve out of outliers. Buckets, percentages
# TODO make a "Golden" list that actually performed well (High over 3%)?

(workbook_path,
beginning_buy_range,
end_buy_range,
beginning_sell_range,
end_sell_range) = Helpers.Collect_User_Input()

num_cols = beginning_sell_range - end_sell_range
num_rows = beginning_buy_range - end_buy_range


# TODO see if you can close the workbook if it is already open
# ? Don't think we can close it unless it only works on windows or linux

try:
    workbook = load_workbook(workbook_path)

except FileNotFoundError:
    print("That workbook doesn't exist. Creating workbook")
    workbook = openpyxl.Workbook()
    workbook.save(workbook_path)

# List of tickers to make
stock_symbols = ["AAPL", "ADBE"]#,"T","WBA","IP","PRU","PM","NEM","F","HAS","LYB","PNW","D","NRG","KEY","VFC","TFC","AAP","IVZ","BBY","AAPL","ADSK","AVGO","CSCO","HPQ","IBM","INTC","INTU","KLAC","MCHP","MRVL","MSFT","NVDA","ORCL","QCOM","TXN","V","BCE","CMCSA","DIS","IPG","LUMN","OMC","T","TM","VIV","VOD","VZ","BBY","DIS","F","GPS","GRMN","HD","LOW","MCD","MAR","NKE","RCL","RL","SBUX","TGT","WMT","YUM","BGS","CAG","CL","CLX","COST","CPB","CVS","DEO","EL","FLO","GIS","HSY","JJSF","K","KDP","KO","KMB","KR","MDLZ","MKC","MO","PEP","SJM","TAP","TSN","WBA","WMT","APA","CNX","CPG","CVE","D","DVN","EGN","ENLC","ENLK","EOG","EQNR","EQT","HES","HP","KMI","MPC","MRO","OKE","OXY","PSX","PTEN","SLB","SM","SU","SWN","VLO","XOM","AFL","AMP","AON","BAC","BEN","BK","BLK","BX","C","CINF","CMA","CME","CNO","COF","DFS","EQR","FITB","GS","HIG","HST","IVZ","JPM","L","LNC","MA","MET","MMC","MTB","MS","NTRS","PNC","PRU","RE","RF","SCHW","STT","TROW","TRV","UNM","USB","V","WFC","ZION","ABT","AMGN","BAX","BMY","CAH","COO","CVS","DHR","JNJ","MCK","MDT","MRK","PFE","RMD","SYK","TFX","TMO","UNH","WBA","WST","XLV","ZBH","AME","AVY","BA","CAT","CSL","DE","DOV","EFX","ECL","EMR","ETN","FDX","FLS","GE","GD","GWW","HON","ITW","LHX","LMT","MMM","MAS","NSC","NOC","PCAR","PH","PNR","ROP","RTX","SEE","SLB","TXT","UNP","UPS","WAB","WM","XYL","AA","ALB","APD","AVD","AVNT","CE","DD","ECL","EMN","FCX","FMC","FNV","GOLD","HUN","IFF","LIN","MLM","MOS","NEU","NEM","NUE","PPG","RPM","SCCO","SHW","SQM","SXT","TG","VALE","VMC","WPM","X","AHT","AKR","AMT","ARE","ARR","AVB","AVD","BXP","CLDT","CMCT","CUZ","CXW","DHC","DLR","EARN","ECL","EPR","EQR","ESS","EXR","FR","GOOD","HT","IRM","IRT","KIM","KRG","LAND","LXP","LSI","LTC","MAA","MLM","MPW","NEM","NEU","NNN","NLY","NHI","NUE","NYMT","O","OLP","PCH","PEAK","PLD","PPG","PSA","REG","RHP","SBRA","SCCO","SLG","SPG","SRC","SUI","STWD","UDR","UMH","VNO","VMC","VTR","WELL","WPC","WY","AES","AVA","BEP","BIP","CMS","CNP","D","DUK","ED","EIX","ETR","EXC","FE","IDA","LNT","NEE","NI","NRG","ORA","PEG","PCG","PNM","PNW","SO","SRE","SJI","TAC","WEC","XEL"]
stock_list = []

# Create all stock objects now
for symbol in stock_symbols:
    stock = Stock(symbol)
    stock_list.append(stock)
        

with open('../logs.txt', 'w') as f:

    # Just to help know how much time is left on the script
    keep_count = 1

    # Loop through all stocks listed above
    # We use a copy because we may need to altar the actual list during the loop
    for stock in stock_list.copy():
    
    # We want to make the ticker used by yfinance to get the dividend dates
    # We use .download() to save all the info about the stock in one variable
        try:
            current_stock_ticker = yf.Ticker(stock.symbol)
            stock_data = yf.download(stock.symbol)
        except ValueError:
            print("Could not find data for symbol: " + stock.symbol)
            continue

        dividends = current_stock_ticker.dividends
        # Get all of the ex-dividend dates
        # Change to python list of timezone-naive objects
        ex_dividend_dates = dividends.index.tz_localize(None).to_pydatetime()

        # The market closes on the weekends so there is no data for the weekends
        # We want to fill in data for the weekends using the next day the market is open
        weekends = pd.date_range(start=stock_data.index.min(), end=stock_data.index.max(), freq='W-SAT')
        weekends = weekends.union(pd.date_range(start=stock_data.index.min(), end=stock_data.index.max(), freq='W-SUN'))
        # Match the format to our data
        weekend_data = stock_data.reindex(weekends)
        # Combine with our data
        stock_data = stock_data.combine_first(weekend_data)
        # Fill the NaN weekend data with the next date we have in the future
        stock_data = stock_data.bfill()

        # Used to know if we should just skip the stock and not add to our Data file
        enough_data = True
        for buy_day in range(end_buy_range, beginning_buy_range + 1):

            for sell_day in range(end_sell_range, beginning_sell_range + 1):
                data_points = 0
                low_outliers = 0
                high_outliers = 0

                for date in ex_dividend_dates:

                    buy_date_dt = datetime.timedelta(days=buy_day)
                    sell_date_dt = datetime.timedelta(days=sell_day)
                    one_day_dt = datetime.timedelta(days=1)

                    
                    # Always us the 'Open' just in case the buy/sell is on a weekend                    
                    try:
                        buy_date_price = stock_data.loc[date - buy_date_dt, 'Open']
                    except:
                        print(stock.symbol, "       No data for this buy day ", buy_date_dt, "   Skipping", file=f)
                        continue
                    try:
                        sell_date_price = stock_data.loc[date - sell_date_dt, 'Open']
                    except:
                        print(stock.symbol, "       No data for this sell day ", sell_date_dt, "    Skipping", file=f)
                        continue
                    
                    # Successfully got buy/sell prices, so we will get a data point
                    data_points += 1

                    price_difference = sell_date_price - buy_date_price
                    stock.single_price_change.append(price_difference)
                    
                    if buy_date_price == 0:
                        print("Tried to divide by 0. Skipping this data point for stock: ", stock.symbol, file=f)
                        data_points -= 1
                        continue
                    #    ((Sell - buy)/buy) * 100
                    percent_change = (((sell_date_price - buy_date_price)/buy_date_price) * 100)
                    stock.single_perc_change.append(percent_change)

                    if percent_change < -3:
                        low_outliers += 1
                    elif percent_change > 5:
                        high_outliers += 1

                # Done with this date range

                if data_points < 50:
                    print("Less than 50 data points for stock:   ", stock.symbol, file=f)
                    enough_data = False
                    break   

                stock.calculate_avg_single_price_change()
                stock.calculate_avg_single_perc_change()
                stock.low_perc_outliers.append(low_outliers)
                stock.high_perc_outliers.append(high_outliers)
                stock.standard_deviation()
                stock.skewness()
                stock.kurtosis()
                stock.clean_single_data()
            
            if enough_data == False:
                break

        print("Done with stock: ", stock.symbol)
        print((keep_count / len(stock_symbols)) * 100, " o/o complete.     just did   ", keep_count)
        keep_count = keep_count + 1


        if enough_data == False:
            # Not enough Data for a date range in a stock, so skip this whole stock and continue to the next
            continue

        
        Helpers.Create_Excel_Raw_Data(workbook, stock, num_cols, num_rows)
        workbook.save(workbook_path)

            

    print("fully done with going through the date ranges")


    print("wow", file=f)
    os.system(f'start excel.exe {workbook_path}')

























    # TODO Getting Data to Excel has changed (see Create_Excel_Raw_Data) but we will need to create the tabels out of that information and the code below could help with that
        # # We are done with the stock, so save its data to the Excel doc
        # work_sheet = workbook.create_sheet(stock.symbol)

        # # Create row titles
        # row_titles = []
        # for num in range(end_buy_range, beginning_buy_range + 1):
        #     row_titles.append(str(num))

        # # Create Column titles
        # col_titles = []
        # for num in range(end_sell_range, beginning_sell_range + 1):
        #     col_titles.append(str(num))



        # # Main data. Averages for each date range
        # for i, title in enumerate(row_titles):
        #     work_sheet.cell(row=i+2, column=1, value=title).font = Font(bold=True)

        # for i, title in enumerate(col_titles):
        #     work_sheet.cell(row=1, column=i+2, value=title).font = Font(bold=True)

        # # Write the data to the worksheet
        # row = 2
        # col = 2
        # for value in stock.multiple_perc_change:
        #     work_sheet.cell(row=row, column=col, value=value)
        #     col += 1
        #     if col >  (2 + beginning_sell_range - end_sell_range):
        #         col = 2
        #         row += 1


        # for i, title in enumerate(row_titles):
        #     work_sheet.cell(row=i+1+beginning_buy_range, column=1, value=title).font = Font(bold=True)
        #     work_sheet.cell(row=i+1+beginning_buy_range, column=(5+(beginning_sell_range - end_sell_range)), value=title).font = Font(bold=True)

        # for i, title in enumerate(col_titles):
        #     work_sheet.cell(row=beginning_buy_range, column=i+2, value=title).font = Font(bold=True)
        #     work_sheet.cell(row=beginning_buy_range, column=i+6+(beginning_sell_range - end_sell_range), value=title).font = Font(bold=True)

        
        # # write the data to the worksheet
        # row = 1+beginning_buy_range
        # col = 2
        # for value in stock.low_perc_outliers:
        #     work_sheet.cell(row=row, column=col, value=value)
        #     col += 1
        #     if col > (2 + (beginning_sell_range - end_sell_range)):
        #         col = 2
        #         row += 1


        # # write the data to the worksheet
        # row = 1+beginning_buy_range
        # col = 6+(beginning_sell_range - end_sell_range)
        # for value in stock.high_perc_outliers:
        #     work_sheet.cell(row=row, column=col, value=value)
        #     col += 1
        #     if col > (5+(beginning_sell_range - end_sell_range)) + (1 + (beginning_sell_range - end_sell_range)):
        #         col = 6+(beginning_sell_range - end_sell_range)
        #         row += 1






    # TODO save things like highest/lowest values for the stock
    # TODO do whatever happens below before saving to excel to add the data
    # final_avg_price_change = []
    # final_avg_perc_change = []
    # first = 0
    # for stock in stock_list:
    #     print("Each day-range percent average for ", stock.symbol, file=f)

    #     if first == 0:
    #         first = 1
    #         final_avg_price_change = stock.multiple_price_change.copy()
    #         final_avg_perc_change = stock.multiple_perc_change.copy()

    #         # for i in range(len(final_avg_perc_change)):
    #         #     print("date-range ", i, "  :   ", stock.multiple_perc_change[i], file=f)
            
    #         continue

    #     for i in range(len(final_avg_price_change)):
    #         #add here, divide by length in the next loop
    #         final_avg_price_change[i] = (final_avg_price_change[i] + stock.multiple_price_change[i])
    #         final_avg_perc_change[i] = (final_avg_perc_change[i] + stock.multiple_perc_change[i])
    #         # print("date-range ", i, "  :   ", stock.multiple_perc_change[i], file=f)



    # for i in range(len(final_avg_price_change)):

    #     final_avg_price_change[i] = final_avg_price_change[i] / len(stock_list)  #this finds the average
    #     final_avg_perc_change[i] = final_avg_perc_change[i] / len(stock_list)


    #     print("Date range ", i, file=f)
    #     print("Average Price change for all stocks in this date range: ", final_avg_price_change[i], file=f)
    #     print("Average Percentage change for all stocks in this date range: ", final_avg_perc_change[i], "%", file=f)