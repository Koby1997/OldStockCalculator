import yfinance as yf
import pandas as pd
import time
import datetime
import openpyxl
from openpyxl import load_workbook
import os
import time
import psutil

from openpyxl.worksheet.datavalidation import DataValidation


pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', 15)

workbook_name = input("What workbook would you like to work on?\n")
workbook_path = f"../Excel_Sheets/{workbook_name}.xlsx"


try:
    workbook = load_workbook(workbook_path)

except FileNotFoundError:
    print("That workbook doesn't exist. Creating workbook")
    workbook = openpyxl.Workbook()
    workbook.save(workbook_path)


# TODO Fix this junk. Try to make a pivot table


def filter_data(stock):
    # Get the columns for the selected stock
    stock_columns = df[stock]

    
    # Drop the first two rows (stock name and calculation type)
    filtered_df = stock_columns.iloc[2:]
    
    # Reset the index and column labels
    filtered_df = filtered_df.reset_index(drop=True)
    filtered_df.columns = ['Average Percent Change', 'High Outliers', 'Low Outliers', 'Standard Deviation', 'Skewness', 'Kurtosis']
    
    return filtered_df

def update_dropdown_input(stock_names):
    # Set the options for the dropdown input
    validation.formula1 = '"{}"'.format('", "'.join(stock_names))
    
    # Save the Excel file
    writer.save()

def write_filtered_data(stock):
    # Filter the data for the selected stock
    filtered_df = filter_data(stock)
    
    # Write the data to a new sheet
    workbook_sheet = workbook[sheet_name]
    filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)
    workbook[sheet_name] = workbook_sheet

# Load the Excel file into a pandas dataframe
df = pd.read_excel(workbook_path, header=1, engine='openpyxl')

# Create a new Excel file and a sheet to hold the dropdown input
writer = pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a')
workbook = load_workbook(workbook_path)

# Create a dropdown input using openpyxl and add it to the new sheet
stock_names = df.columns.get_level_values(0).unique().tolist()

validation = DataValidation(type="list", formula1='"{}"'.format('", "'.join(stock_names)))


worksheet = workbook.create_sheet('Dropdown Input')
worksheet.add_data_validation(validation)
validation.add('A1')

# Write the filtered data to a new sheet for each stock
for stock_name in stock_names:
    sheet_name = f"{stock_name} Filtered Data"
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    write_filtered_data(stock_name)

# Update the dropdown input options
update_dropdown_input(stock_names)

# Close the Excel writer object
writer.save()
writer.close()



os.system(f'start excel.exe {workbook_path}')
