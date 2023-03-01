# import pandas as pd

# # create a dictionary of data
# data = {'name': ['Alice', 'Bob', 'Charlie', 'Dave'],
#         'age': [25, 32, 18, 47],
#         'country': ['USA', 'Canada', 'France', 'UK']}

# # create a pandas DataFrame from the data
# df = pd.DataFrame(data)

# # save the DataFrame to an Excel file
# df.to_excel('test.xlsx', index=False)



from openpyxl import Workbook
from openpyxl import load_workbook
import os

# create a new workbook and select the active worksheet
wb = load_workbook('testTwo.xlsx')
ws = wb.active
ws2 = wb.create_sheet("This new name")

# create a list of data
data = [1, 2, 3, 4, 5,
        6, 7, 8, 9, 10,
        1111111, 12, 13, 14, 15,
        16, 17, 18, 19, 20,
        21, 22, 23, 24, 25]

# write the data to the worksheet
row = 1
col = 1
for value in data:
    ws2.cell(row=5+row, column=5+col, value=value)
    col += 1
    if col > 5:
        col = 1
        row += 1

# save the workbook to a file
wb.save('testTwo.xlsx')
os.system('start excel.exe "testTwo.xlsx"')
